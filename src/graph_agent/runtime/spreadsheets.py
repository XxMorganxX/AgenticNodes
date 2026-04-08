from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
import csv
from pathlib import Path
from typing import Any

from graph_agent.runtime.run_documents import normalize_run_documents

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency guard
    load_workbook = None


SUPPORTED_SPREADSHEET_FORMATS = {"csv", "xlsx"}
SPREADSHEET_STORAGE_SUFFIXES = {".csv", ".xlsx"}
DEFAULT_SAMPLE_ROW_LIMIT = 5
SPREADSHEET_HEADER_ROW_INDEX = 1
SPREADSHEET_FIRST_DATA_ROW_INDEX = 2


class SpreadsheetParseError(ValueError):
    """Raised when a spreadsheet file cannot be parsed with the current config."""


def resolve_spreadsheet_path_from_run_documents(
    documents: Any,
    *,
    run_document_id: str = "",
    run_document_name: str = "",
) -> str:
    """Pick storage_path from run-attached documents when file_path is unset."""
    normalized = normalize_run_documents(documents)
    ready: list[dict[str, Any]] = []
    for doc in normalized:
        if str(doc.get("status") or "") != "ready":
            continue
        path = str(doc.get("storage_path") or "").strip()
        if not path:
            continue
        if Path(path).suffix.lower() not in SPREADSHEET_STORAGE_SUFFIXES:
            continue
        ready.append(doc)
    if not ready:
        return ""
    doc_id = str(run_document_id or "").strip()
    if doc_id:
        for doc in ready:
            if str(doc.get("document_id") or "") == doc_id:
                return str(doc.get("storage_path") or "").strip()
        return ""
    name = str(run_document_name or "").strip()
    if name:
        for doc in ready:
            doc_name = str(doc.get("name") or "")
            if doc_name == name or doc_name.lower() == name.lower():
                return str(doc.get("storage_path") or "").strip()
        return ""
    if len(ready) == 1:
        return str(ready[0].get("storage_path") or "").strip()
    return ""


@dataclass(frozen=True)
class SpreadsheetRowRecord:
    row_number: int
    row_data: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "row_data": dict(self.row_data),
        }


@dataclass(frozen=True)
class SpreadsheetParseResult:
    source_file: str
    file_format: str
    sheet_name: str | None
    sheet_names: list[str]
    headers: list[str]
    rows: list[SpreadsheetRowRecord]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def preview(self, *, limit: int = DEFAULT_SAMPLE_ROW_LIMIT) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "file_format": self.file_format,
            "sheet_name": self.sheet_name,
            "sheet_names": list(self.sheet_names),
            "headers": list(self.headers),
            "row_count": self.row_count,
            "sample_rows": [row.to_dict() for row in self.rows[: max(0, int(limit))]],
        }


@dataclass(frozen=True)
class SpreadsheetMatrixRow:
    row_label: str
    row_number: int
    values: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_label": self.row_label,
            "row_number": self.row_number,
            "values": dict(self.values),
        }


@dataclass(frozen=True)
class SpreadsheetMatrixParseResult:
    source_file: str
    file_format: str
    sheet_name: str | None
    sheet_names: list[str]
    corner_label: str
    column_labels: list[str]
    rows: list[SpreadsheetMatrixRow]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.column_labels)

    @property
    def row_labels(self) -> list[str]:
        return [row.row_label for row in self.rows]

    def row_by_label(self, row_label: str) -> SpreadsheetMatrixRow:
        normalized = str(row_label or "").strip()
        for row in self.rows:
            if row.row_label == normalized:
                return row
        raise KeyError(f"Unknown matrix row label '{normalized}'.")

    def column_number_for_label(self, column_label: str) -> int:
        normalized = str(column_label or "").strip()
        for index, label in enumerate(self.column_labels, start=2):
            if label == normalized:
                return index
        raise KeyError(f"Unknown matrix column label '{normalized}'.")

    def cell_value(self, row_label: str, column_label: str) -> Any:
        row = self.row_by_label(row_label)
        return row.values[column_label]

    def preview(self, *, limit: int = DEFAULT_SAMPLE_ROW_LIMIT) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "file_format": self.file_format,
            "sheet_name": self.sheet_name,
            "sheet_names": list(self.sheet_names),
            "corner_label": self.corner_label,
            "column_labels": list(self.column_labels),
            "row_labels": self.row_labels,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "sample_rows": [row.to_dict() for row in self.rows[: max(0, int(limit))]],
        }


@dataclass(frozen=True)
class SpreadsheetGridData:
    source_file: str
    file_format: str
    sheet_name: str | None
    sheet_names: list[str]
    raw_rows: list[list[Any]]


def infer_spreadsheet_format(file_path: str, file_format: str | None = None) -> str:
    normalized = str(file_format or "").strip().lower()
    if normalized == "auto":
        normalized = ""
    if normalized in SUPPORTED_SPREADSHEET_FORMATS:
        return normalized
    suffix = Path(file_path).suffix.lower().lstrip(".")
    if suffix in SUPPORTED_SPREADSHEET_FORMATS:
        return suffix
    raise SpreadsheetParseError("Spreadsheet format must be one of: csv, xlsx, or auto-detected from the file extension.")


def parse_spreadsheet(
    *,
    file_path: str,
    file_format: str | None = None,
    sheet_name: str | None = None,
    header_row_index: int = 1,
    start_row_index: int | None = None,
    empty_row_policy: str = "skip",
) -> SpreadsheetParseResult:
    grid = _load_spreadsheet_grid(file_path=file_path, file_format=file_format, sheet_name=sheet_name)
    normalized_empty_policy = str(empty_row_policy or "skip").strip().lower()
    if normalized_empty_policy not in {"skip", "include"}:
        raise SpreadsheetParseError("Empty row policy must be either 'skip' or 'include'.")
    header_row_index = SPREADSHEET_HEADER_ROW_INDEX
    start_row_index = SPREADSHEET_FIRST_DATA_ROW_INDEX
    return _build_parse_result(
        source_file=grid.source_file,
        file_format=grid.file_format,
        sheet_name=grid.sheet_name,
        sheet_names=grid.sheet_names,
        raw_rows=grid.raw_rows,
        header_row_index=header_row_index,
        start_row_index=start_row_index,
        empty_row_policy=normalized_empty_policy,
    )


def parse_spreadsheet_matrix(
    *,
    file_path: str,
    file_format: str | None = None,
    sheet_name: str | None = None,
) -> SpreadsheetMatrixParseResult:
    grid = _load_spreadsheet_grid(file_path=file_path, file_format=file_format, sheet_name=sheet_name)
    return _build_matrix_parse_result(
        source_file=grid.source_file,
        file_format=grid.file_format,
        sheet_name=grid.sheet_name,
        sheet_names=grid.sheet_names,
        raw_rows=grid.raw_rows,
    )


def _load_spreadsheet_grid(
    *,
    file_path: str,
    file_format: str | None = None,
    sheet_name: str | None = None,
) -> SpreadsheetGridData:
    normalized_path = str(file_path).strip()
    if not normalized_path:
        raise SpreadsheetParseError(
            "Spreadsheet file path is required. Set file_path on the Spreadsheet Rows node, choose a ready project "
            "file, attach exactly one ready CSV/XLSX run document, or set run_document_id / run_document_name to "
            "choose among several."
        )
    path = Path(normalized_path).expanduser()
    if not path.exists() or not path.is_file():
        raise SpreadsheetParseError(f"Spreadsheet file not found: {normalized_path}")

    resolved_format = infer_spreadsheet_format(normalized_path, file_format)
    if resolved_format == "csv":
        return SpreadsheetGridData(
            source_file=str(path),
            file_format="csv",
            sheet_name=None,
            sheet_names=[],
            raw_rows=_read_csv_rows(path),
        )
    selected_sheet_name, sheet_names, raw_rows = _read_xlsx_rows(path, sheet_name=sheet_name)
    return SpreadsheetGridData(
        source_file=str(path),
        file_format="xlsx",
        sheet_name=selected_sheet_name,
        sheet_names=sheet_names,
        raw_rows=raw_rows,
    )


def _read_csv_rows(path: Path) -> list[list[Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def _read_xlsx_rows(path: Path, *, sheet_name: str | None) -> tuple[str, list[str], list[list[Any]]]:
    if load_workbook is None:
        raise SpreadsheetParseError("XLSX support requires the 'openpyxl' package to be installed.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    selected_sheet_name = str(sheet_name or "").strip()
    if selected_sheet_name:
        if selected_sheet_name not in workbook.sheetnames:
            raise SpreadsheetParseError(
                f"Sheet '{selected_sheet_name}' was not found. Available sheets: {', '.join(sheet_names) or 'none'}."
            )
        worksheet = workbook[selected_sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]
        selected_sheet_name = worksheet.title
    return selected_sheet_name, sheet_names, [list(row) for row in worksheet.iter_rows(values_only=True)]


def _build_parse_result(
    *,
    source_file: str,
    file_format: str,
    sheet_name: str | None,
    sheet_names: list[str],
    raw_rows: Sequence[Sequence[Any]],
    header_row_index: int,
    start_row_index: int | None,
    empty_row_policy: str,
) -> SpreadsheetParseResult:
    if header_row_index > len(raw_rows):
        raise SpreadsheetParseError(
            f"Header row {header_row_index} is outside the available range of {len(raw_rows)} row(s)."
        )
    header_values = list(raw_rows[header_row_index - 1]) if raw_rows else []
    headers = _normalize_headers(header_values)
    first_data_row = start_row_index if start_row_index is not None else header_row_index + 1
    if first_data_row <= header_row_index:
        raise SpreadsheetParseError("First data row must come after the header row.")
    row_records: list[SpreadsheetRowRecord] = []
    for row_number in range(first_data_row, len(raw_rows) + 1):
        row_values = list(raw_rows[row_number - 1])
        normalized_values = [_normalize_cell_value(value) for value in row_values]
        row_data = {
            header: normalized_values[index] if index < len(normalized_values) else None
            for index, header in enumerate(headers)
        }
        if empty_row_policy == "skip" and _row_is_empty(row_data.values()):
            continue
        row_records.append(
            SpreadsheetRowRecord(
                row_number=row_number,
                row_data=row_data,
            )
        )
    return SpreadsheetParseResult(
        source_file=source_file,
        file_format=file_format,
        sheet_name=sheet_name,
        sheet_names=sheet_names,
        headers=headers,
        rows=row_records,
    )


def _build_matrix_parse_result(
    *,
    source_file: str,
    file_format: str,
    sheet_name: str | None,
    sheet_names: list[str],
    raw_rows: Sequence[Sequence[Any]],
) -> SpreadsheetMatrixParseResult:
    if len(raw_rows) < 2:
        raise SpreadsheetParseError(
            "Spreadsheet matrix must include a first-row column axis and at least one labeled data row."
        )
    header_values = list(raw_rows[0])
    if len(header_values) < 2:
        raise SpreadsheetParseError(
            "Spreadsheet matrix must include a first-column row axis and at least one decision column."
        )
    if any(len(row) > len(header_values) for row in raw_rows[1:]):
        raise SpreadsheetParseError(
            "Spreadsheet matrix contains data columns beyond the first-row headers. Add labels for every populated column."
        )

    corner_label = _stringify_axis_label(header_values[0])
    column_labels = [
        _stringify_axis_label(value, fallback=f"column_{index}")
        for index, value in enumerate(header_values[1:], start=2)
    ]
    _raise_for_duplicate_axis_labels(column_labels, axis_name="column")

    rows: list[SpreadsheetMatrixRow] = []
    seen_row_labels: set[str] = set()
    for row_number in range(2, len(raw_rows) + 1):
        raw_row = list(raw_rows[row_number - 1])
        row_label = _stringify_axis_label(raw_row[0] if raw_row else None)
        cell_values = [_normalize_cell_value(value) for value in raw_row[1:]]
        if not row_label and _row_is_empty(cell_values):
            continue
        if not row_label:
            raise SpreadsheetParseError(f"Spreadsheet matrix row {row_number} is missing a first-column row label.")
        if row_label in seen_row_labels:
            raise SpreadsheetParseError(f"Spreadsheet matrix has duplicate row label '{row_label}'.")
        seen_row_labels.add(row_label)
        values = {
            column_label: cell_values[index] if index < len(cell_values) else None
            for index, column_label in enumerate(column_labels)
        }
        rows.append(
            SpreadsheetMatrixRow(
                row_label=row_label,
                row_number=row_number,
                values=values,
            )
        )

    if not rows:
        raise SpreadsheetParseError("Spreadsheet matrix must include at least one labeled data row.")

    return SpreadsheetMatrixParseResult(
        source_file=source_file,
        file_format=file_format,
        sheet_name=sheet_name,
        sheet_names=sheet_names,
        corner_label=corner_label,
        column_labels=column_labels,
        rows=rows,
    )


def _normalize_headers(values: Sequence[Any]) -> list[str]:
    if not values:
        raise SpreadsheetParseError("Header row is empty.")
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        candidate = str(value or "").strip()
        if not candidate:
            candidate = f"column_{index}"
        candidate = candidate.replace("\n", " ").replace("\r", " ").strip()
        candidate = "_".join(part for part in candidate.split(" ") if part)
        normalized = candidate or f"column_{index}"
        count = seen.get(normalized, 0) + 1
        seen[normalized] = count
        headers.append(normalized if count == 1 else f"{normalized}_{count}")
    return headers


def _stringify_axis_label(value: Any, *, fallback: str = "") -> str:
    if value is None:
        return fallback
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    rendered = str(value).replace("\n", " ").replace("\r", " ").strip()
    return rendered or fallback


def _raise_for_duplicate_axis_labels(labels: Sequence[str], *, axis_name: str) -> None:
    seen: set[str] = set()
    for label in labels:
        if label in seen:
            raise SpreadsheetParseError(f"Spreadsheet matrix has duplicate {axis_name} label '{label}'.")
        seen.add(label)


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _row_is_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True
